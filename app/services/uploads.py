from __future__ import annotations

import csv
from io import BytesIO, StringIO

import openpyxl

from app.schemas import JobItemRequest


WORKUA_URL_HEADERS = ("workua_url", "my-0 href")
COMPANY_HEADERS = ("company_name", "strong-600 2")


def parse_upload_bytes(filename: str, payload: bytes) -> list[JobItemRequest]:
    lower_name = filename.lower()
    if lower_name.endswith(".xlsx"):
        return _parse_xlsx(payload)
    if lower_name.endswith(".csv"):
        return _parse_csv(payload)
    raise ValueError("Unsupported file type. Use .xlsx or .csv")


def _parse_xlsx(payload: bytes) -> list[JobItemRequest]:
    workbook = openpyxl.load_workbook(BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    url_idx = _resolve_header(headers, WORKUA_URL_HEADERS)
    company_idx = _resolve_header(headers, COMPANY_HEADERS, required=False)
    items: list[JobItemRequest] = []
    for row_index in range(2, sheet.max_row + 1):
        workua_url = sheet.cell(row_index, url_idx + 1).value
        if not workua_url:
            continue
        company_name = sheet.cell(row_index, company_idx + 1).value if company_idx is not None else ""
        items.append(
            JobItemRequest(
                row_index=row_index,
                company_name=str(company_name or ""),
                workua_url=str(workua_url).strip(),
            )
        )
    return items


def _parse_csv(payload: bytes) -> list[JobItemRequest]:
    reader = csv.DictReader(StringIO(payload.decode("utf-8-sig")))
    items: list[JobItemRequest] = []
    for row_index, row in enumerate(reader, start=2):
        workua_url = (row.get("workua_url") or "").strip()
        if not workua_url:
            continue
        items.append(
            JobItemRequest(
                row_index=row_index,
                company_name=(row.get("company_name") or "").strip(),
                workua_url=workua_url,
            )
        )
    return items


def _resolve_header(headers: list[object], candidates: tuple[str, ...], *, required: bool = True) -> int | None:
    normalized = {str(header).strip().lower(): idx for idx, header in enumerate(headers) if header is not None}
    for candidate in candidates:
        if candidate.lower() in normalized:
            return normalized[candidate.lower()]
    if required:
        raise ValueError(f"Missing required header. Expected one of: {', '.join(candidates)}")
    return None
